import openpyxl

wb = openpyxl.Workbook()

print(wb.create_sheet())

wb.create_sheet(index=0, title='First Sheet')

wb.create_sheet(index=2, title='Middle Sheet')

del wb['Middle Sheet']
del wb['Sheet1']
print(wb.sheetnames)