import openpyxl

wb = openpyxl.Workbook()

sheet = wb.active
sheet.title = 'Spam Bacon Eggs Sheet'
print(wb.sheetnames)
print(sheet.title)
print(wb.sheetnames)