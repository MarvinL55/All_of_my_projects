import openpyxl

wb = openpyxl.load_workbook("C:\\Users\\marvi\\Downloads\\updatedProduceSales.xlsx")
sheet = wb['Sheet']

PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.10,
                 'Lemon': 1.27}

for rowNum in range(2, sheet.max_row):
    productName = sheet.cell(row=rowNum, column=1).value
    if productName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[productName]

wb.save('C:\\Users\\marvi\\Downloads\\updatedProduceSales.xlsx')